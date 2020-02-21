import openpyxl
from openpyxl import *
from openpyxl.styles import *
import os
import pyexcel as p
from copy import copy
from openpyxl.styles.borders import Border, Side
import sys

def formatAR(passing_offset, wb_sheet):
    #name / matter number / whatveer that other number is
    offset = passing_offset
    cord = 'A' + str((offset+4))
    final_sheet[cord] = wb_sheet['B1'].value + '.' + wb_sheet['B2'].value + ' - ' + wb_sheet['C1'].value + ' / ' + wb_sheet['C2'].value
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True)

    cord = 'A' + str((offset+5))
    final_sheet[cord] = "A/R LEDGER HISTORY"
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True)

    cord = 'A' + str((offset+7))
    final_sheet[cord] = 'Bill #'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'left')

    cord = 'B' + str((offset+7))
    final_sheet[cord] = 'Date'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')

    cord = 'C' + str((offset+7))
    final_sheet[cord] = 'Payment'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'right')

    cord = 'D' + str((offset+7))
    final_sheet[cord] = "Fees"
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'right')

    cord = 'E' + str((offset+7))
    final_sheet[cord] = 'Expenses'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'right')

    cord = 'F' + str((offset+7))
    final_sheet[cord] = 'Balance'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'right')

    cord = 'H' + str((offset+7))
    final_sheet[cord] = 'Comment'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')


    #add 8 to offset to get proper position



    # Bill numbers
    for row in wb_sheet.iter_rows(min_row=5, min_col = 2, max_col = 2, max_row=1000):
        for cell in row:
            if(cell.value is None):
                break
            cord = "A" + str((cell.row + offset+3))
            num = int(cell.value)
            final_sheet[cord] = num
            final_sheet[cord].font = Font(name = 'arial', size = 10)
            final_sheet[cord].alignment = Alignment(horizontal = 'left')
            new_offset = int(cell.row + offset + 3)

    # Dates
    for row in wb_sheet.iter_rows(min_row=5, min_col = 3, max_col = 3, max_row=1000):
        for cell2 in row:
            if(cell2.value is None):
                break
            cord = "B" + str(cell2.row + offset + 3)
            date = str(cell2.value)
            #remove the time from the date 
            final_sheet[cord] = date.split(" ",1)[0]
            final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #payments
    for row in wb_sheet.iter_rows(min_row=5, min_col = 4, max_col = 4, max_row=1000):
        for cell3 in row:
            if(cell3.value is None):
                break
            cord = "C" + str(cell3.row + offset+3)
            Payment = float(cell3.value)
            final_sheet[cord] = Payment
            final_sheet[cord].alignment = Alignment(horizontal = 'right')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #Fees
    for row in wb_sheet.iter_rows(min_row=5, min_col = 5, max_col = 5, max_row=1000):
        for cell4 in row:
            if(cell4.value is None):
                break
            cord = "D" + str(cell4.row + offset+3)
            fee = float(cell4.value)
            final_sheet[cord] = fee
            final_sheet[cord].alignment = Alignment(horizontal = 'right')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #expenses
    for row in wb_sheet.iter_rows(min_row=5, min_col = 6, max_col = 6, max_row=1000):
        for cell5 in row:
            if(cell5.value is None):
                break
            cord = "E" + str(cell5.row + offset+3)
            expense= float(cell5.value)
            final_sheet[cord] = expense
            final_sheet[cord].alignment = Alignment(horizontal = 'right')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #Balance
    for row in wb_sheet.iter_rows(min_row=5, min_col = 8, max_col = 8, max_row=1000):
        for cell6 in row:
            if(cell6.value is None):
                break
            cord = "F" + str(cell6.row + offset+3)
            balance= float(cell6.value)
            final_sheet[cord] = balance
            final_sheet[cord].alignment = Alignment(horizontal = 'right')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #Comments
    for row in wb_sheet.iter_rows(min_row=5, min_col = 9, max_col = 9, max_row=1000):
        for cell7 in row:
            if(cell7.value is None):
                break
            cord = "H" + str(cell7.row + offset+3)
            comments= str(cell7.value)
            final_sheet[cord] = comments
            final_sheet[cord].alignment = Alignment(horizontal = 'right')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #totals
    cord = 'A' + str((new_offset+1))
    final_sheet[cord] = "TOTALS"
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True)
    final_sheet[cord].alignment = Alignment(horizontal = 'left')


    # Border between
    for i in range(0,8):
        cord = abc[i] + str(new_offset+2)
        final_sheet[cord].border = Border(bottom=Side(style='thin'))

    return new_offset

def formatTL(passing_offset, wb_sheet):
    
    offset = passing_offset
    
    cord = 'A' + str((offset+4))
    final_sheet[cord] = "TRUST ACCOUNT LEDGER HISTORY"
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True)

    cord = 'B' + str((offset+6))
    final_sheet[cord] = 'Date'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')

    cord = 'C' + str((offset+6))
    final_sheet[cord] = 'Type'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')

    cord = 'D' + str((offset+6))
    final_sheet[cord] = 'Check#'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')

    cord = 'E' + str((offset+6))
    final_sheet[cord] = 'Amount'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'right')

    cord = 'F' + str((offset+6))
    final_sheet[cord] = 'Balance'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'right')

    cord = 'h' + str((offset+7))
    final_sheet[cord] = 'Memo'
    final_sheet[cord].font = Font(name = 'arial', size = 10, bold = True, underline = 'single')
    final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')

    #date
    for row in wb_sheet.iter_rows(min_row=5, min_col = 3, max_col = 3, max_row=1000):
        for cell2 in row:
            if(cell2.value is None):
                break
            cord = "B" + str(cell2.row + offset + 2)
            date = str(cell2.value)
            #remove the time from the date 
            final_sheet[cord] = date.split(" ",1)[0]
            final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #type
    for row in wb_sheet.iter_rows(min_row=5, min_col = 4, max_col = 4, max_row=1000):
        for cell2 in row:
            if(cell2.value is None):
                break
            cord = "C" + str(cell2.row + offset + 2)
            theType = str(cell2.value)
            final_sheet[cord] = theType
            final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')
            final_sheet[cord].font = Font(name = 'arial', size = 10)
    
    #check number
    for row in wb_sheet.iter_rows(min_row=5, min_col = 5, max_col = 5, max_row=1000):
        for cell2 in row:
            if(cell2.value is None):
                break
            cord = "D" + str(cell2.row + offset + 2)
            checkNum = str(cell2.value)
            final_sheet[cord] = checkNum
            final_sheet[cord].alignment = Alignment(horizontal = 'centerContinuous')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #Amount
    for row in wb_sheet.iter_rows(min_row=5, min_col = 6, max_col = 6, max_row=1000):
        for cell2 in row:
            if(cell2.value is None):
                break
            cord = "E" + str(cell2.row + offset + 2)
            theType = float(cell2.value)
            final_sheet[cord] = theType
            final_sheet[cord].alignment = Alignment(horizontal = 'right')
            final_sheet[cord].font = Font(name = 'arial', size = 10)

    #Balance
    for row in wb_sheet.iter_rows(min_row=5, min_col = 7, max_col = 7, max_row=1000):
        for cell2 in row:
            if(cell2.value is None):
                break
            cord = "F" + str(cell2.row + offset + 2)
            Bal = float(cell2.value)
            final_sheet[cord] = Bal
            final_sheet[cord].alignment = Alignment(horizontal = 'right')
            final_sheet[cord].font = Font(name = 'arial', size = 10)
    #memo
    for row in wb_sheet.iter_rows(min_row=5, min_col = 8, max_col = 8, max_row=1000):
        for cell2 in row:
            if(cell2.value is None):
                break
            cord = "H" + str(cell2.row + offset + 2)
            memo = str(cell2.value)
            final_sheet[cord] = memo
            final_sheet[cord].alignment = Alignment(horizontal = 'left')
            final_sheet[cord].font = Font(name = 'arial', size = 10)


#create a new workbook

final_wb = Workbook()
final_sheet = final_wb.active
final_sheet.title = "Client-Matter Inquiry"


#set all cells to height 15 ( kind of sloppy not sure how else to do it)
abc = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
for i in range(0,1000):
    final_sheet.row_dimensions[i].height = 15

for i in range(0,26):
    final_sheet.column_dimensions[abc[i]].width = 15


##start adding AR ledgers
passing_offset = 0
userinputbook =''
ledgerHistoryInputBook = ''

print("-------------------------------------------------------------------------")
print("### ATTENTION #### - ALL FILES  NEED to be in the same folder this program is located in. #### ")
print("--------------------------------------------------------------")
print("Type exact name of A/R ledger history file, WITH its extension .xls (it is an .xls which is old. Will create new copy to work with that is new .xlsx format) Then press ENTER")
print("OR type 'finished' all lower case when done adding AR ledgers")
print("-------------------------------------------------------------------------")
userinputbook = input()

while(userinputbook != 'finished'):

    #convert from xls to xlsx
    newname = userinputbook.split(".",1)[0] + "_fixed" + ".xlsx"
    p.save_book_as(file_name=userinputbook, dest_file_name = newname)
    print("........")
    print("......")
    print("...")
    print(".")
    print("-------------------------------------------------------------------------")
    print("created new file from the given xls to a new xlsx new name = ", newname)
    print("-------------------------------------------------------------------------")
    print(".")
    print("...")
    print("......")
    print("........")
    wb = openpyxl.load_workbook(newname)
    wb_sheet = wb.get_sheet_by_name('Client-Matter Inquiry')
    #My formatAR function
    returnedoffset = formatAR(passing_offset, wb_sheet)
    passing_offset = returnedoffset

    #reloop or exit prompt user for new input
    print("-------------------------------------------------------------------------")
    print("To add ANOTHER A/R file Type exact name of A/R ledger history file, with its extension, Then press ENTER")
    print("OR type 'finished' all lower case when done adding AR ledgers")
    print("-------------------------------------------------------------------------")
    userinputbook = input()

print("........")
print("......")
print("...")
print(".")
print("-------------------------------------------------------------------------")
print("You typed 'finished' so we are Done adding AR files. Now to add an optional optional history file ")
print("Type exact name Trust Ledger file WITH its extension (it is an .xls which is old. Will create new copy to work with that is new .xlsx format) Then press ENTER (Example - testbook.xls) (file should be in this folder)")
print("OR type 'skip' to skip this part and press ENTER")
print("-------------------------------------------------------------------------")


#add trust account ledger history

userinputbook = input()

if(userinputbook != "skip"):
    #convert from xls to xlsx
    newname = userinputbook.split(".",1)[0] + "fixed" + ".xlsx"
    p.save_book_as(file_name=userinputbook, dest_file_name = newname)
    print("........")
    print("......")
    print("...")
    print(".")
    print("-------------------------------------------------------------------------")
    print("created new file from the given xls to a new xlsx new name = ", newname)
    print("-------------------------------------------------------------------------")
    print(".")
    print("...")
    print("......")
    print("........")
    wb = openpyxl.load_workbook(newname)
    wb_sheet = wb.get_sheet_by_name('Client-Matter Inquiry')

    #My formatHistory function
    formatTL(passing_offset, wb_sheet)
    
    #save new file and exit program
    print("........")
    print("......")
    print("...")
    print(".")
    print("-------------------------------------------------------------------------")
    print("You added Trust history ledger, Time to save the file and be done")
    print("Type the name you want the final file to be saved as Then press ENTER ")
    print(" !!!DO NOT include its extension!!!! (.xlsx)")
    print("-------------------------------------------------------------------------")
    finalname = input()
    final_wb.save(finalname + ".xlsx")
    print("........")
    print("......")
    print("...")
    print(".")
    print("-------------------------------------------------------------------------")
    print("saved file as -> ", finalname, ".xlsx")
    print("-------------------------------------------------------------------------")
    print(".")
    print("...")
    print("......")
    print("........")
    print("-------------------------------------------------------------------------")
    print(" ALL DONE! have fun slacking off now mom !")
    print("-------------------------------------------------------------------------")
    
else:
    #save new file and exit program without adding a history

    print("........")
    print("......")
    print("...")
    print(".")
    print("-------------------------------------------------------------------------")
    print("You typed 'skip' Time to save the file and be done")
    print("Type the name you want the final file to be saved as Then press ENTER ")
    print(" !!!DO NOT include its extension!!!! (.xlsx)")
    print("-------------------------------------------------------------------------")
    finalname = input()
    final_wb.save(finalname + ".xlsx")
    print("........")
    print("......")
    print("...")
    print(".")
    print("-------------------------------------------------------------------------")
    print("saved file as -> ", finalname,".xlsx")
    print("-------------------------------------------------------------------------")
    print(".")
    print("...")
    print("......")
    print("........")
    print("-------------------------------------------------------------------------")
    print(" ALL DONE! have fun slacking off now mom !")
    print("-------------------------------------------------------------------------")









